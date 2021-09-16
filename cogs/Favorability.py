# 필수 임포트
from discord.ext import commands
import discord
import os
from utils import logger

# 부가 임포트
from utils import util_box
import openpyxl
import random
random.random()

class FavorabilityCommand(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    @commands.command()
    async def 학습(self, ctx):
        word = ctx.split(" ")
        word_ = ctx.replace("그웬아 학습 ", "").replace(word[1], "")
        wb = openpyxl.load_workbook("memory.xlsx")
        ws = wb.active

        for i in range(1, 10):
            if ws["A" + str(i)].value == "" or ws["A" + str(i)].value == word[1]:
                ws["A" + str(i)].value = word[1]
                ws["B" + str(i)].value = word_
                await ctx.send(f'알았어, {ctx.author.name}!')
                break
        wb.save("memory.xlsx")

def setup(bot):
    logger.info(f'{os.path.abspath(__file__)} 로드 완료')
    bot.add_cog(FavorabilityCommand(bot)) 
