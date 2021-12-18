from discord.ext import commands
import discord
import os
from utils import logger

from utils import util_box
import openpyxl
import random
random.random()


class EventCog(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
    
    @commands.Cog.listener()
    async def on_message(self, message):
        if message.author.bot:  # 봇이 말한 건 무시
            return

        logger.msg(message)  # 메시지를 기록

        if message.content.endswith("바보"):
            await message.channel.send('바보')

        if message.content.startswith("그웬아") or message.content.startswith("!"):
            word = message.content.split(" ")
            wb = openpyxl.load_workbook("memory.xlsx")
            ws = wb.active

            for i in range(1, 10):
                if ws["A" + str(i)].value == word[1]:
                    await message.channel.send(ws["B" + str(i)].value)
                    break

    @commands.Cog.listener()
    async def on_member_join(self, member):
        pass

    @commands.Cog.listener()
    async def on_member_remove(self, member):
        pass

    @commands.Cog.listener()
    async def on_guild_join(self, guild):
        logger.info(f"{guild.name} 서버에 들어갔어!")

    @commands.Cog.listener()
    async def on_guild_remove(self, guild):
        logger.info(f"{guild.name} 서버에서 쫓겨났어...")

    @commands.Cog.listener()
    async def on_message_delete(self, message):
        if message.author.bot:
            return
        logger.info(f"{message.author.name}이(가) '{message.content}' 메시지를 삭제")

    @commands.Cog.listener()
    async def on_message_edit(self, before, after):
        if before.author.bot:
            return
        logger.info(f"{before.author.name}이(가) '{before.content}' 메시지를 '{after.content}' 로 수정")

    @commands.Cog.listener()
    async def on_command_error(self, ctx, error):
        if isinstance(error, discord.errors.Forbidden):  # 권한 부족 오류
            try:
                return await ctx.send('내 권한이 부족해.. ㅠㅠ')
            except discord.errors.Forbidden:
                return logger.warn('채팅 권한 없음')

        elif isinstance(error, commands.errors.CommandNotFound):  # 해당하는 명령어가 없는 경우
            return await ctx.send('몰루, 히힛!')

        elif isinstance(error, commands.CommandOnCooldown):  # 명령어 쿨타임이 다 차지 않은 경우
            return await ctx.send(
                f'이 명령어는 {error.cooldown.rate}번 쓰면 {error.cooldown.per}초의 쿨타임이 생기는 명령어야!'
                f'``cs\n{int(error.retry_after)}초 남았어!``')

        await ctx.send(f'오류가 발생했어...\n`{str(error)}`')
        logger.err(error)


def setup(bot):
    logger.info(f'{os.path.abspath(__file__)} 로드 완료')
    bot.add_cog(EventCog(bot))
